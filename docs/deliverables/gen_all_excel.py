#!/usr/bin/env python3
# SOFIA Documentation Agent — gen_all_excel.py
# Genera Quality Dashboard + Sprint Metrics + Velocity Report para Sprints 4-8
# Ejecutar: python3 gen_all_excel.py  (desde docs/deliverables/)
# Prerequisito: bash .sofia/scripts/install-deps.sh

import sys, os
from pathlib import Path

# ── Auto-detectar Python con openpyxl (venv SOFIA o sistema) ─────────────────
def _bootstrap_openpyxl():
    try:
        import openpyxl
        return
    except ImportError:
        pass
    candidates = [
        Path.home() / '.sofia-venv' / 'bin' / 'python3',
        Path(os.environ.get('REPO_ROOT', '')) / '.sofia' / 'venv' / 'bin' / 'python3',
    ]
    for py in candidates:
        if py.exists():
            os.execv(str(py), [str(py)] + sys.argv)
    print("❌ openpyxl no encontrado. Ejecuta: bash .sofia/scripts/install-deps.sh")
    sys.exit(1)

_bootstrap_openpyxl()

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

# ── Ruta base ─────────────────────────────────────────────────────────────────
BASE = Path(__file__).parent

# ── Paleta Experis ────────────────────────────────────────────────────────────
BLUE='1B3A6B'; MED='2E5F9E'; VL='EBF3FB'; GPASS='C6EFCE'; GDONE='E2EFDA'
RED='FFCCCC'; YEL='FFEB9C'; WHITE='FFFFFF'; GRAY='CCCCCC'

def side(): return Side(style='thin', color=GRAY)
def brd():  return Border(left=side(),right=side(),top=side(),bottom=side())
def hfont(sz=11): return Font(name='Arial',bold=True,color=WHITE,size=sz)
def dfont(bold=False,color='000000',sz=10): return Font(name='Arial',size=sz,bold=bold,color=color)
def hfill(): return PatternFill('solid',fgColor=BLUE)
def afill(): return PatternFill('solid',fgColor=VL)
def wfill(): return PatternFill('solid',fgColor=WHITE)
def sfill(s):
    u=str(s).upper()
    if any(k in u for k in ('PASS','OK','DONE','APPROVED','CUMPLE','CERRADO','✅')): return PatternFill('solid',fgColor=GPASS)
    if any(k in u for k in ('FAIL','ERROR','❌')): return PatternFill('solid',fgColor=RED)
    if any(k in u for k in ('BLOCKED','OPEN','WARN','ABIERTO','PARCIAL','⚠')): return PatternFill('solid',fgColor=YEL)
    return wfill()
def cc(): return Alignment(horizontal='center',vertical='center',wrap_text=True)
def lc(): return Alignment(horizontal='left',vertical='top',wrap_text=True)
def ap(cell,font=None,fill=None,align=None):
    if font:  cell.font=font
    if fill:  cell.fill=fill
    if align: cell.alignment=align
    cell.border=brd()
def title(ws,row,text,ncols):
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=ncols)
    c=ws.cell(row=row,column=1,value=text)
    ap(c,hfont(13),hfill(),cc()); ws.row_dimensions[row].height=26
def hrow(ws,row,cols):
    for i,(text,width) in enumerate(cols,1):
        c=ws.cell(row=row,column=i,value=text)
        ap(c,hfont(),hfill(),cc()); ws.column_dimensions[get_column_letter(i)].width=width
    ws.row_dimensions[row].height=22
def drow(ws,row,vals,alt=False,scols=None):
    f=afill() if alt else wfill()
    for i,v in enumerate(vals,1):
        c=ws.cell(row=row,column=i,value=v)
        use=sfill(v) if scols and i in scols else f
        ap(c,dfont(),use,lc()); ws.row_dimensions[row].height=18
def total_row(ws,row,vals,fills):
    for i,(v,fill) in enumerate(zip(vals,fills),1):
        c=ws.cell(row=row,column=i,value=v)
        c.font=Font(name='Arial',bold=True,size=11,color=WHITE if fill==BLUE else '000000')
        c.fill=PatternFill('solid',fgColor=fill); c.alignment=cc(); c.border=brd()

# ── Sprint Data ───────────────────────────────────────────────────────────────
SPRINTS = {
    4: {'feat':'FEAT-003','titulo':'Dispositivos de Confianza','version':'v1.3.0','fecha':'2026-05-16','dir':'sprint-4-FEAT-003','sp':24,
        'items':[('US-201','Gestionar dispositivos de confianza',4,'new-feature','DONE'),('US-202','Agregar dispositivo + OTP',5,'new-feature','DONE'),('US-203','Revocar dispositivo individual o todos',4,'new-feature','DONE'),('US-204','Expiración automática 90 días',4,'new-feature','DONE'),('US-205','Notificaciones de cambio',4,'new-feature','DONE'),('DEBT-004','ua-parser-java fingerprint',2,'tech-debt','DONE'),('ACT-Flyway','Flyway V6 trusted_devices',1,'infra','DONE')],
        'areas':[('US-201 — Gestión dispositivos',6,6,'100%','PASS'),('US-202 — Agregar + OTP',8,8,'100%','PASS'),('US-203 — Revocar',6,6,'100%','PASS'),('US-204 — Expiración 90d',5,5,'100%','PASS'),('US-205 — Notificaciones',5,5,'100%','PASS'),('DEBT-004 — ua-parser',3,3,'100%','PASS'),('Seguridad + E2E',5,5,'100%','PASS')],
        'tcs':38,'pci':[('8.6.1','Dispositivos gestionados y auditados','CUMPLE'),('10.2.7','Cambios config. autenticación','CUMPLE')],
        'kpis':[('TCs acumulados Sprints 1-4','175'),('Defectos totales','0'),('PCI-DSS req. 8.6.1','CUMPLE'),('Gates HITL acum.','20')]},
    5: {'feat':'FEAT-005','titulo':'Panel Auditoría (inicio)','version':'v1.4.0','fecha':'2026-06-06','dir':'sprint-5-FEAT-005','sp':24,
        'items':[('US-401','Dashboard seguridad resumen visual',4,'new-feature','DONE'),('US-402','Exportar historial PDF y CSV',3,'new-feature','DONE'),('DEBT-005','CompletableFuture.allOf() 6 queries',3,'tech-debt','DONE'),('DEBT-006','OpenPDF async streaming',2,'tech-debt','DONE'),('DEBT-007','ADR-010 SSE CORS CDN',3,'tech-debt','DONE'),('ACT-HLD','HLD FEAT-005 + C4',2,'documental','DONE'),('ACT-LLD','LLD-005 backend + frontend',2,'documental','DONE'),('ACT-ADR','ADR-009 HMAC key rotation',1,'documental','DONE'),('ACT-OpenAPI','OpenAPI v1.4.0',1,'documental','DONE'),('ACT-E2E','E2E dashboard + export',3,'testing','DONE')],
        'areas':[('US-401 — Security Dashboard',8,8,'100%','PASS'),('US-402 — Exportación PDF/CSV',7,7,'100%','PASS'),('DEBT-005 — CompletableFuture',4,4,'100%','PASS'),('DEBT-006 — OpenPDF async',3,3,'100%','PASS'),('DEBT-007 — SSE config',3,3,'100%','PASS'),('E2E Playwright dashboard',5,5,'100%','PASS')],
        'tcs':35,'pci':[('10.7','Historial consultable por titular','CUMPLE'),('10.2.1','Eventos registrados','CUMPLE'),('10.3.3','Exportación con SHA-256','CUMPLE')],
        'kpis':[('TCs acumulados Sprints 1-5','210'),('Defectos totales','0'),('PCI-DSS req. 10.7','CUMPLE'),('Gates HITL acum.','25')]},
    6: {'feat':'FEAT-005+004','titulo':'Preferencias + Notificaciones inicio','version':'v1.5.0','fecha':'2026-06-20','dir':'sprint-6-FEAT-005','sp':24,
        'items':[('US-403','Preferencias seguridad unificadas',3,'new-feature','DONE'),('US-301','Historial paginado notificaciones 90d',4,'new-feature','DONE'),('US-302','Mark-as-read individual + bulk',2,'new-feature','DONE'),('US-303','Badge unread-count @Cacheable 30s',3,'new-feature','DONE'),('DEBT-008','CompletableFuture.allOf() optimización',3,'tech-debt','DONE'),('Flyway-V7','Flyway V7 user_notifications',2,'infra','DONE'),('ADR-010/011','ADR-010 SSE + ADR-011 JWT context',2,'documental','DONE'),('LLD-006','LLD-006 NotificationCenter',2,'documental','DONE'),('OpenAPI-1.4.1','OpenAPI v1.4.1',1,'documental','DONE'),('ACT-E2E','E2E preferencias + notificaciones',2,'testing','DONE')],
        'areas':[('US-403 — Preferencias seguridad',6,6,'100%','PASS'),('US-301/302/303 inicio',8,8,'100%','PASS'),('DEBT-008 — Dashboard paralelo',5,5,'100%','PASS'),('Flyway V7 + infra',3,3,'100%','PASS'),('Documentación ADR/LLD',4,4,'100%','PASS')],
        'tcs':32,'pci':[('8.6.3','Preferencias autenticación gestionadas','CUMPLE'),('10.2.1','Cambios config registrados','CUMPLE')],
        'kpis':[('TCs acumulados Sprints 1-6','242'),('Defectos totales','0'),('Velocidad media','24 SP/sprint'),('Gates HITL acum.','30')]},
    7: {'feat':'FEAT-006','titulo':'Autenticación Contextual y Bloqueo','version':'v1.7.0','fecha':'2026-06-23','dir':'sprint-7-FEAT-006','sp':24,
        'items':[('US-601','Bloqueo automático ventana 24h',5,'new-feature','DONE'),('US-602','Desbloqueo email HMAC TTL 1h',3,'new-feature','DONE'),('US-603','Login contextual context-pending ADR-011',5,'new-feature','DONE'),('US-604','Historial config 90d PCI-DSS 10.2',4,'new-feature','DONE'),('US-403','Preferencias seguridad (FEAT-005 cierre)',3,'new-feature','DONE'),('DEBT-008','SecurityDashboard CompletableFuture',3,'tech-debt','DONE'),('ACT-30','OpenAPI v1.4.0 claims JWT',1,'documental','DONE')],
        'areas':[('US-601 — Bloqueo cuenta',8,8,'100%','PASS'),('US-602 — Desbloqueo email',6,6,'100%','PASS'),('US-603 — Login contextual',9,9,'100%','PASS'),('US-604 — Historial config',6,6,'100%','PASS'),('US-403 — Preferencias',4,4,'100%','PASS'),('DEBT-008 — Dashboard',4,4,'100%','PASS')],
        'tcs':81,'pci':[('8.3.4','Bloqueo por intentos fallidos (US-601)','CUMPLE'),('8.3.4','Desbloqueo proceso controlado (US-602)','CUMPLE'),('10.2','Historial cambios configuración (US-604)','CUMPLE')],
        'kpis':[('TCs Sprint 7','81'),('Defectos totales','0'),('PCI-DSS req. 8.3.4 + 10.2','CUMPLE'),('Gates HITL Sprint 7','5/5')]},
    8: {'feat':'FEAT-004','titulo':'Centro de Notificaciones de Seguridad','version':'v1.8.0','fecha':'2026-03-17','dir':'sprint-8-FEAT-004','sp':24,
        'items':[('US-301','Historial paginado 90d + filtros',4,'new-feature','DONE'),('US-302','Mark-as-read + SSE broadcast',2,'new-feature','DONE'),('US-303','Badge @Cacheable 30s + SSE inicial',3,'new-feature','DONE'),('US-304','Deep-links + revoke session',4,'new-feature','DONE'),('US-305','SSE 1 conn/usuario + heartbeat + fallback',3,'new-feature','DONE'),('DEBT-009','JWT blacklist Redis TTL=JWT-restante',3,'tech-debt','DONE'),('DEBT-010','extractIpSubnet /24 centralizado',2,'tech-debt','DONE'),('Flyway-V9','V9 user_notifications + índices',2,'infra','DONE'),('ACT-31','OpenAPI v1.5.0 FEAT-004 + SSE',1,'documental','DONE')],
        'areas':[('US-301 — Historial paginado',5,5,'100%','PASS'),('US-302 — Mark-as-read',6,6,'100%','PASS'),('US-303 — Badge count',3,3,'100%','PASS'),('US-304 — Acciones directas',9,9,'100%','PASS'),('US-305 — SSE + Angular',6,6,'100%','PASS'),('DEBT-009 — JWT Blacklist',6,6,'100%','PASS'),('SseRegistry ADR-012',8,8,'100%','PASS'),('E2E Playwright 13 escenarios',13,13,'100%','PASS')],
        'tcs':59,'pci':[('8.3','JWT context-pending blacklisted tras uso único','CUMPLE'),('10.2.1','Auditoría NOTIFICATIONS_VIEWED por consulta','CUMPLE')],
        'kpis':[('TCs Sprint 8','59'),('Defectos totales','0'),('ADR-012 SSE pool','200 conn, 1/usuario'),('Gates HITL Sprint 8','6/6')]},
}

ALL_SPRINTS = [
    ('Sprint 1','2026-03-02 — 2026-03-16',24,24,'FEAT-001 2FA TOTP'),
    ('Sprint 2','2026-03-30 — 2026-04-10',24,24,'FEAT-001 2FA cierre'),
    ('Sprint 3','2026-04-14 — 2026-04-25',24,24,'FEAT-002 Sesiones'),
    ('Sprint 4','2026-05-05 — 2026-05-16',24,24,'FEAT-003 Dispositivos'),
    ('Sprint 5','2026-05-26 — 2026-06-06',24,24,'FEAT-005 Auditoría'),
    ('Sprint 6','2026-06-09 — 2026-06-20',24,24,'FEAT-005+004'),
    ('Sprint 7','2026-06-09 — 2026-06-23',24,24,'FEAT-006 Auth Contextual'),
    ('Sprint 8','2026-03-03 — 2026-03-17',24,24,'FEAT-004 Notificaciones'),
]

def get_dir(snum):
    d = BASE / SPRINTS[snum]['dir'] / 'excel'
    d.mkdir(parents=True, exist_ok=True)
    return d

def build_quality_dashboard(snum):
    s = SPRINTS[snum]; d = get_dir(snum)
    wb = Workbook(); ws = wb.active; ws.title = 'Resumen'
    ws.sheet_view.showGridLines = False
    title(ws,1,f"Quality Dashboard — Sprint {snum} — BankPortal — {s['feat']}",5)
    ws.merge_cells('A2:E2'); c=ws.cell(row=2,column=1,value=f"{s['feat']} | SOFIA QA Agent | {s['fecha']} | {s['version']}")
    c.font=Font(name='Arial',size=10,italic=True,color='444444'); c.fill=PatternFill('solid',fgColor=VL); c.alignment=cc()
    hrow(ws,3,[('Área',28),('TCs Plan',12),('TCs Real',12),('Cobertura',12),('Estado',12)])
    for i,area in enumerate(s['areas']): drow(ws,i+4,list(area),i%2==1,{5})
    tr=len(s['areas'])+4
    total_row(ws,tr,[f"TOTAL Sprint {snum}",s['tcs'],s['tcs'],'100%','PASS'],[BLUE,GPASS,GPASS,GPASS,GPASS])
    kr=tr+2; ws.merge_cells(start_row=kr,start_column=1,end_row=kr,end_column=5)
    c=ws.cell(row=kr,column=1,value='KPIs'); c.font=hfont(); c.fill=hfill(); c.alignment=cc()
    for j,(k,v) in enumerate(s['kpis']):
        r=kr+1+j; c1=ws.cell(row=r,column=1,value=k); ap(c1,dfont(True),afill() if j%2==0 else wfill(),lc())
        ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=3)
        c2=ws.cell(row=r,column=4,value=v); ap(c2,Font(name='Arial',bold=True,size=11),PatternFill('solid',fgColor=GPASS),cc())
        ws.merge_cells(start_row=r,start_column=4,end_row=r,end_column=5)
    ws2=wb.create_sheet('Items Sprint'); ws2.sheet_view.showGridLines=False
    title(ws2,1,f"Items Sprint {snum} — Estado",5)
    hrow(ws2,2,[('ID',12),('Título',42),('SP',8),('Tipo',14),('Estado',12)])
    for i,item in enumerate(s['items']): drow(ws2,i+3,list(item),i%2==1,{5})
    total_row(ws2,len(s['items'])+3,[f"TOTAL",f"{len(s['items'])} items",s['sp'],'','DONE'],[BLUE,BLUE,BLUE,BLUE,GPASS])
    ws3=wb.create_sheet('PCI-DSS 4.0'); ws3.sheet_view.showGridLines=False
    title(ws3,1,f"Checklist PCI-DSS 4.0 — Sprint {snum}",4)
    hrow(ws3,2,[('Requisito',12),('Descripción',42),('Verificación',32),('Estado',12)])
    for i,row in enumerate(s['pci']): drow(ws3,i+3,[row[0],'',row[1]],i%2==1,{3})
    fpath = d/f"Quality-Dashboard-Sprint{snum}.xlsx"; wb.save(fpath); print(f'  ✅ {fpath}')

def build_sprint_metrics(snum):
    s = SPRINTS[snum]; d = get_dir(snum)
    wb = Workbook(); ws = wb.active; ws.title = 'Resumen Sprint'
    ws.sheet_view.showGridLines = False
    title(ws,1,f"Sprint Metrics — Sprint {snum} — {s['feat']}",5)
    ws.merge_cells('A2:E2'); c=ws.cell(row=2,column=1,value=f"Sprint Goal ALCANZADO | {s['fecha']} | {s['version']}")
    c.font=Font(name='Arial',size=10,italic=True); c.fill=PatternFill('solid',fgColor=VL); c.alignment=cc()
    hrow(ws,3,[('Métrica',28),('Planificado',14),('Real',14),('Variación',14),('Estado',14)])
    us_count = len([x for x in s['items'] if x[3]=='new-feature'])
    debt_count = len([x for x in s['items'] if x[3]=='tech-debt'])
    rows=[('Story Points',24,24,'0','OK'),('US completadas',us_count,us_count,'0','OK'),('DEBT resueltas',debt_count,debt_count,'0','OK'),('Tests nuevos','-',s['tcs'],'-','OK'),('Defectos QA',0,0,'0','OK')]
    for i,(m,p,r,v,st) in enumerate(rows):
        row=i+4; c1=ws.cell(row=row,column=1,value=m); ap(c1,dfont(True),afill() if i%2==1 else wfill(),lc())
        for col,val in [(2,p),(3,r),(4,v)]:
            c=ws.cell(row=row,column=col,value=val); ap(c,dfont(),afill() if i%2==1 else wfill(),cc())
        cs=ws.cell(row=row,column=5,value=st); ap(cs,Font(name='Arial',size=10,bold=True),sfill(st),cc())
    sr=len(rows)+4; ws.merge_cells(start_row=sr,start_column=1,end_row=sr,end_column=3)
    c=ws.cell(row=sr,column=1,value=f'VELOCIDAD SPRINT {snum}'); c.font=Font(name='Arial',bold=True,size=12,color=WHITE); c.fill=hfill(); c.alignment=cc()
    ws.merge_cells(start_row=sr,start_column=4,end_row=sr,end_column=5)
    c2=ws.cell(row=sr,column=4,value='24 SP / sprint'); c2.font=Font(name='Arial',bold=True,size=12); c2.fill=PatternFill('solid',fgColor=GPASS); c2.alignment=cc(); c2.border=brd()
    ws2=wb.create_sheet('Items Detalle'); ws2.sheet_view.showGridLines=False
    title(ws2,1,f"Items Sprint {snum} — Detalle",5)
    hrow(ws2,2,[('ID',12),('Título',42),('SP',8),('Tipo',14),('Estado',12)])
    for i,item in enumerate(s['items']): drow(ws2,i+3,list(item),i%2==1,{5})
    total_row(ws2,len(s['items'])+3,['TOTAL',f"{len(s['items'])} items",s['sp'],'',f"✅ {s['sp']}/{s['sp']} SP"],[BLUE,BLUE,BLUE,BLUE,GPASS])
    fpath = d/f"Sprint-Metrics-Sprint{snum}.xlsx"; wb.save(fpath); print(f'  ✅ {fpath}')

def build_velocity_report():
    d = get_dir(8)
    wb = Workbook(); ws = wb.active; ws.title = 'Velocidad'
    ws.sheet_view.showGridLines = False
    title(ws,1,'Velocity Report — BankPortal — 8 Sprints Completados',6)
    ws.merge_cells('A2:F2'); c=ws.cell(row=2,column=1,value='BankPortal · Banco Meridian · 8 Sprints · Velocidad media: 24 SP/sprint · 0 defectos')
    c.font=Font(name='Arial',size=10,italic=True); c.fill=PatternFill('solid',fgColor=VL); c.alignment=cc()
    hrow(ws,3,[('Sprint',14),('Período',24),('SP Plan',12),('SP Real',12),('Feature',28),('Estado',14)])
    for i,(sp,per,plan,real,feat) in enumerate(ALL_SPRINTS):
        row=i+4
        for col,val,aln in [(1,sp,cc()),(2,per,lc()),(3,plan,cc()),(4,real,cc()),(5,feat,lc())]:
            c=ws.cell(row=row,column=col,value=val); ap(c,dfont(col==1),afill() if i%2==1 else wfill(),aln)
        cv=ws.cell(row=row,column=6,value='✅ DONE'); ap(cv,Font(name='Arial',bold=True,size=10),PatternFill('solid',fgColor=GPASS),cc())
    ar=len(ALL_SPRINTS)+4; ws.merge_cells(start_row=ar,start_column=1,end_row=ar,end_column=2)
    c=ws.cell(row=ar,column=1,value='VELOCIDAD MEDIA — 8 SPRINTS'); c.font=Font(name='Arial',bold=True,color=WHITE); c.fill=hfill(); c.alignment=cc(); c.border=brd()
    for col,val in [(3,'=AVERAGE(C4:C11)'),(4,'=AVERAGE(D4:D11)'),(5,'192 SP acumulados'),(6,'0 defectos')]:
        c=ws.cell(row=ar,column=col,value=val); ap(c,Font(name='Arial',bold=True),PatternFill('solid',fgColor=GPASS),cc())
    chart=BarChart(); chart.type='col'; chart.title='Velocidad por Sprint — BankPortal 8 Sprints'
    chart.y_axis.title='Story Points'; chart.style=10; chart.width=22; chart.height=12
    chart.add_data(Reference(ws,min_col=4,min_row=3,max_row=ar-1),titles_from_data=True)
    chart.set_categories(Reference(ws,min_col=1,min_row=4,max_row=ar-1)); ws.add_chart(chart,'H3')
    ws2=wb.create_sheet('SP Acumulados'); ws2.sheet_view.showGridLines=False
    title(ws2,1,'Story Points Acumulados — 8 Sprints',4)
    hrow(ws2,2,[('Sprint',14),('SP Sprint',14),('SP Acumulado',18),('Feature',32)])
    acum=0
    for i,(sp,_,plan,_,feat) in enumerate(ALL_SPRINTS):
        acum+=plan; row=i+3
        for col,val,aln in [(1,sp,cc()),(2,plan,cc()),(3,acum,cc()),(4,feat,lc())]:
            c=ws2.cell(row=row,column=col,value=val); ap(c,dfont(col==1),afill() if i%2==1 else wfill(),aln)
    total_row(ws2,len(ALL_SPRINTS)+3,['TOTAL',192,192,'8 Sprints · 23.875 SP/sprint media'],[BLUE,GPASS,GPASS,GPASS])
    fpath = d/'Velocity-Report-Completo.xlsx'; wb.save(fpath); print(f'  ✅ {fpath}')

if __name__ == '__main__':
    print('\n📊 SOFIA Documentation Agent — Generando Excel sprints 4-8...\n')
    for snum in [4,5,6,7,8]:
        s = SPRINTS[snum]
        print(f'── Sprint {snum} ({s["feat"]}) ──')
        build_quality_dashboard(snum)
        build_sprint_metrics(snum)
    print('\n── Velocity Report global (8 sprints) ──')
    build_velocity_report()
    print('\n✅ TODOS los Excel generados correctamente.\n')
