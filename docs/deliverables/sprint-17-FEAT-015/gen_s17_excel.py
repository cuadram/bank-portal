import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from openpyxl.utils import get_column_letter
import os

OUT = os.path.join(os.path.dirname(__file__), 'excel')

BLUE    = 'FF1B3A6B'
BLUE2   = 'FF2E5F9E'
WHITE   = 'FFFFFFFF'
LGRAY   = 'FFF5F7FA'
MGRAY   = 'FFE2E8F0'
GREEN   = 'FF16A34A'
LGREEN  = 'FFE8F5E9'
YELLOW  = 'FFFFF9C4'
RED     = 'FFDC2626'
LRED    = 'FFFFEBEE'

def hdr_font(color='FFFFFFFF', bold=True, size=11): return Font(name='Arial', bold=bold, color=color, size=size)
def hdr_fill(color): return PatternFill('solid', fgColor=color)
def thin_border():
    s = Side(style='thin', color='FFCCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)
def center(): return Alignment(horizontal='center', vertical='center', wrap_text=True)
def left(): return Alignment(horizontal='left', vertical='center', wrap_text=True)

def style_hdr(cell, bg=BLUE, fg='FFFFFFFF', bold=True):
    cell.font = hdr_font(fg, bold)
    cell.fill = hdr_fill(bg)
    cell.alignment = center()
    cell.border = thin_border()

def style_cell(cell, bold=False, bg=WHITE, fg='FF000000', align='left'):
    cell.font = Font(name='Arial', size=10, bold=bold, color=fg)
    cell.fill = hdr_fill(bg)
    cell.alignment = left() if align=='left' else center()
    cell.border = thin_border()

# ── 1. NC-Tracker-Sprint17.xlsx ──────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'NC Tracker Sprint 17'
ws.sheet_view.showGridLines = False

# Title
ws.merge_cells('A1:I1')
ws['A1'] = 'NC TRACKER — Sprint 17 · FEAT-015 Transferencias Programadas · BankPortal'
ws['A1'].font = Font(name='Arial', bold=True, size=13, color='FFFFFFFF')
ws['A1'].fill = hdr_fill(BLUE)
ws['A1'].alignment = center()

ws.merge_cells('A2:I2')
ws['A2'] = 'CMMI Level 3 · VER SP 3.1 · Total NCs: 3  |  Bloqueantes: 0  |  Resueltas: 3  |  Abiertas: 0'
ws['A2'].font = Font(name='Arial', size=10, color='FF2E5F9E')
ws['A2'].alignment = center()
ws['A2'].fill = hdr_fill(LGRAY)

hdrs = ['ID','Tipo','Descripción','Fichero','Severidad','Reportado por','Estado','Sprint detectado','Resolución']
for col, h in enumerate(hdrs, 1):
    c = ws.cell(row=3, column=col, value=h)
    style_hdr(c)

data = [
    ['NC-017-001','Estilo','Nombre variable no sigue convención camelCase','ScheduledTransferService.java','Menor','Code Reviewer','✅ Resuelta','S17','Renombrada en misma PR'],
    ['NC-017-002','Documentación','Javadoc faltante en método executeScheduled()','ScheduledTransferExecutor.java','Menor','Code Reviewer','✅ Resuelta','S17','Javadoc añadido'],
    ['NC-017-003','Test','Test unitario TC-1503 sin assertion en caso borde','ScheduledTransferExecutorTest.java','Menor','QA Lead','✅ Resuelta','S17','Assertion añadida'],
]
for row, d in enumerate(data, 4):
    for col, val in enumerate(d, 1):
        c = ws.cell(row=row, column=col, value=val)
        bg = LGREEN if '✅' in str(val) else (LRED if val in ['Crítica','Mayor'] else WHITE)
        style_cell(c, bg=bg)
        if col == 1: style_cell(c, bold=True, bg=LGRAY)

cols = [10,10,42,38,10,16,14,16,32]
for i, w in enumerate(cols, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[1].height = 28
ws.row_dimensions[2].height = 20
for r in range(3, 7):
    ws.row_dimensions[r].height = 18

wb.save(os.path.join(OUT, 'NC-Tracker-Sprint17.xlsx'))
print('✅ NC-Tracker-Sprint17.xlsx')

# ── 2. Decision-Log-Sprint17.xlsx ────────────────────────────────────────────
wb2 = openpyxl.Workbook()
ws2 = wb2.active
ws2.title = 'Decision Log Sprint 17'
ws2.sheet_view.showGridLines = False

ws2.merge_cells('A1:H1')
ws2['A1'] = 'DECISION LOG — Sprint 17 · FEAT-015 Transferencias Programadas · BankPortal'
ws2['A1'].font = Font(name='Arial', bold=True, size=13, color='FFFFFFFF')
ws2['A1'].fill = hdr_fill(BLUE)
ws2['A1'].alignment = center()

ws2.merge_cells('A2:H2')
ws2['A2'] = 'CMMI Level 3 · Decision Analysis · Total decisiones: 4'
ws2['A2'].font = Font(name='Arial', size=10, color='FF2E5F9E')
ws2['A2'].alignment = center()
ws2['A2'].fill = hdr_fill(LGRAY)

hdrs2 = ['ID','Fecha','Descripción','Alternativas consideradas','Decisión tomada','Autor','Impacto','Estado']
for col, h in enumerate(hdrs2, 1):
    c = ws2.cell(row=3, column=col, value=h)
    style_hdr(c)

decisions = [
    ['ADR-026','2026-04-08','Scheduler transferencias programadas','(A) Quartz, (B) Spring @Scheduled + idempotency_key','B — Spring @Scheduled + idempotency_key UUID','Tech Lead','Implementación simplificada sin dependencia externa','✅ Implementado'],
    ['ADR-026b','2026-04-10','Tabla DB para transferencias','(A) En tabla transfers, (B) Tabla dedicada scheduled_transfers','B — Tabla dedicada scheduled_transfers','Tech Lead','Separación limpia de dominios, Flyway V17','✅ Implementado'],
    ['SEC-017-01','2026-04-09','Cifrado push subscriptions','(A) Cifrado a nivel app, (B) AES-256-GCM columnas DB','B — AES-256-GCM en columnas auth y p256dh','Tech Lead + Sec','DEBT-028 cerrada, datos protegidos en reposo','✅ Implementado'],
    ['ARCH-017-01','2026-04-15','Recurrencia configurable','(A) Enum fijo DAILY/WEEKLY/MONTHLY, (B) cron expression','A — Enum tipado — suficiente para MVP, extensible','Tech Lead','Menos complejidad, validación más simple','✅ Implementado'],
]
for row, d in enumerate(decisions, 4):
    for col, val in enumerate(d, 1):
        c = ws2.cell(row=row, column=col, value=val)
        bg = LGREEN if '✅' in str(val) else WHITE
        style_cell(c, bg=bg)
        if col == 1: style_cell(c, bold=True, bg=LGRAY, fg=BLUE2)

cols2 = [12,12,38,44,44,12,44,14]
for i, w in enumerate(cols2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.row_dimensions[1].height = 28
for r in range(3, 8):
    ws2.row_dimensions[r].height = 32

wb2.save(os.path.join(OUT, 'Decision-Log-Sprint17.xlsx'))
print('✅ Decision-Log-Sprint17.xlsx')

# ── 3. Quality-Dashboard-Sprint17.xlsx ───────────────────────────────────────
wb3 = openpyxl.Workbook()
ws3 = wb3.active
ws3.title = 'Quality Dashboard'
ws3.sheet_view.showGridLines = False

ws3.merge_cells('A1:F1')
ws3['A1'] = 'QUALITY DASHBOARD — Sprint 17 · FEAT-015 · BankPortal · v1.17.0'
ws3['A1'].font = Font(name='Arial', bold=True, size=14, color='FFFFFFFF')
ws3['A1'].fill = hdr_fill(BLUE)
ws3['A1'].alignment = center()

ws3.merge_cells('A2:F2')
ws3['A2'] = 'CMMI Level 3 · MA SP 1.1/2.1 · Período: 2026-04-08 → 2026-04-22'
ws3['A2'].font = Font(name='Arial', size=10, color='FF2E5F9E', italic=True)
ws3['A2'].alignment = center()
ws3['A2'].fill = hdr_fill(LGRAY)

sections = [
    ('MÉTRICAS DE CALIDAD', [
        ('KPI','Valor S17','Valor S16','Umbral','Delta','Estado'),
        ('Tests QA ejecutados','45','45','≥ 30','=','✅'),
        ('Tests unitarios acumulados','615','553','≥ 500','+62','✅'),
        ('Cobertura application','85%','84%','≥ 80%','+1%','✅'),
        ('Defectos en producción','0','0','= 0','=','✅'),
        ('CVEs críticos/altos','0','0','= 0','=','✅'),
        ('NCs bloqueantes CR','0','0','= 0','=','✅'),
        ('Sprint Goal cumplido','100%','100%','≥ 90%','=','✅'),
    ]),
    ('MÉTRICAS DE PROCESO', [
        ('Métrica','Valor','Referencia','Tendencia','',''),
        ('SP entregados','24','24 SP (S16)','→ Estable','',''),
        ('SP acumulados','401','377 (S16)','+24 SP','',''),
        ('Velocidad media (18 sp→17)','23.6 SP/sprint','23.6','→ Estable','',''),
        ('Deuda técnica cerrada','3 items (DEBT-027/028/029)','2 (S16)','↑ Mejora','',''),
        ('Riesgos cerrados en sprint','2 (R-016-01, R-016-05)','0 (S16)','↑ Mejora','',''),
    ]),
]

current_row = 4
for section_title, rows in sections:
    ws3.merge_cells(f'A{current_row}:F{current_row}')
    ws3[f'A{current_row}'] = section_title
    ws3[f'A{current_row}'].font = Font(name='Arial', bold=True, size=11, color='FFFFFFFF')
    ws3[f'A{current_row}'].fill = hdr_fill(BLUE2)
    ws3[f'A{current_row}'].alignment = center()
    ws3.row_dimensions[current_row].height = 20
    current_row += 1
    for r_idx, row in enumerate(rows):
        for col, val in enumerate(row, 1):
            c = ws3.cell(row=current_row, column=col, value=val)
            if r_idx == 0:
                style_hdr(c, bg=MGRAY, fg='FF1B3A6B')
            else:
                is_ok = str(val) == '✅'
                bg = LGREEN if is_ok else (LGRAY if col == 1 else WHITE)
                style_cell(c, bold=(col==1), bg=bg)
        current_row += 1
    current_row += 1

col_widths = [36, 28, 24, 16, 10, 10]
for i, w in enumerate(col_widths, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

# Second sheet — Sprint history
ws_hist = wb3.create_sheet('Sprint History')
ws_hist.merge_cells('A1:F1')
ws_hist['A1'] = 'Histórico de Sprints — BankPortal (S1 → S17)'
ws_hist['A1'].font = Font(name='Arial', bold=True, size=13, color='FFFFFFFF')
ws_hist['A1'].fill = hdr_fill(BLUE)
ws_hist['A1'].alignment = center()

hist_hdrs = ['Sprint','Feature','SP','Cobertura','Tests','Defectos PRD']
for col, h in enumerate(hist_hdrs, 1):
    c = ws_hist.cell(row=2, column=col, value=h)
    style_hdr(c)

hist = [
    ['S1','FEAT-001 Auth',18,'72%',88,0],
    ['S2','Deuda S1',11,'74%',104,0],
    ['S3','FEAT-002 Dashboard',21,'75%',142,0],
    ['S4','FEAT-003 Transferencias',22,'76%',178,0],
    ['S5-8','FEAT-004/005 Historial/Push',23,'78%',230,0],
    ['S9','Deuda técnica',11,'79%',280,0],
    ['S10','FEAT-008 Seguridad',24,'80%',320,0],
    ['S11','FEAT-009 Reporting',22,'81%',355,0],
    ['S12','FEAT-010 Core Banking',23,'82%',390,0],
    ['S13','FEAT-011 Multi-cuenta',24,'83%',430,0],
    ['S14','FEAT-012 OAuth2/PKCE',22,'83%',468,0],
    ['S15','FEAT-013 Auditoría',22,'84%',500,0],
    ['S16','FEAT-014 Push VAPID',24,'84%',553,0],
    ['S17','FEAT-015 Transferencias Prog.',24,'85%',615,0],
]
for row, d in enumerate(hist, 3):
    for col, val in enumerate(d, 1):
        c = ws_hist.cell(row=row, column=col, value=val)
        style_cell(c, bold=(col==1), bg=LGRAY if col==1 else (LGREEN if col==6 and val==0 else WHITE))

for i, w in enumerate([8,38,8,12,10,14], 1):
    ws_hist.column_dimensions[get_column_letter(i)].width = w

wb3.save(os.path.join(OUT, 'Quality-Dashboard-Sprint17.xlsx'))
print('✅ Quality-Dashboard-Sprint17.xlsx')

print('\n✅ Sprint 17 — 3 ficheros Excel generados en:', OUT)
