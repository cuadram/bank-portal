import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

OUT = os.path.join(os.path.dirname(__file__), 'excel')

BLUE='FF1B3A6B'; BLUE2='FF2E5F9E'; WHITE='FFFFFFFF'; LGRAY='FFF5F7FA'
MGRAY='FFE2E8F0'; LGREEN='FFE8F5E9'; YELLOW='FFFFF9C4'; LRED='FFFFEBEE'

def hf(c='FFFFFFFF',b=True,s=11): return Font(name='Arial',bold=b,color=c,size=s)
def hfill(c): return PatternFill('solid',fgColor=c)
def tbr():
    s=Side(style='thin',color='FFCCCCCC')
    return Border(left=s,right=s,top=s,bottom=s)
def ca(): return Alignment(horizontal='center',vertical='center',wrap_text=True)
def la(): return Alignment(horizontal='left',vertical='center',wrap_text=True)

def sh(c,bg=BLUE,fg='FFFFFFFF'):
    c.font=hf(fg); c.fill=hfill(bg); c.alignment=ca(); c.border=tbr()
def sc(c,bold=False,bg=WHITE,fg='FF000000'):
    c.font=Font(name='Arial',size=10,bold=bold,color=fg)
    c.fill=hfill(bg); c.alignment=la(); c.border=tbr()

# ── NC-Tracker-Sprint18.xlsx ─────────────────────────────────────────────────
wb=openpyxl.Workbook(); ws=wb.active
ws.title='NC Tracker Sprint 18'; ws.sheet_view.showGridLines=False

ws.merge_cells('A1:I1')
ws['A1']='NC TRACKER — Sprint 18 · FEAT-016 Gestión de Tarjetas · BankPortal'
ws['A1'].font=Font(name='Arial',bold=True,size=13,color='FFFFFFFF')
ws['A1'].fill=hfill(BLUE); ws['A1'].alignment=ca()
ws.merge_cells('A2:I2')
ws['A2']='CMMI Level 3 · VER SP 3.1 · Total NCs: 5  |  Bloqueantes: 0  |  Resueltas: 5  |  Abiertas: 0'
ws['A2'].font=Font(name='Arial',size=10,color='FF2E5F9E'); ws['A2'].alignment=ca(); ws['A2'].fill=hfill(LGRAY)

for col,h in enumerate(['ID','Tipo','Descripción','Fichero','Severidad','Reportado por','Estado','Sprint','Resolución'],1):
    c=ws.cell(row=3,column=col,value=h); sh(c)

ncs=[
    ['NC-018-001','Seguridad','Falta @PreAuthorize en CardController.getCards()','CardController.java','Mayor','Security Reviewer','✅ Resuelta','S18','@PreAuthorize("hasRole(USER)") añadido'],
    ['NC-018-002','Seguridad','PAN visible en log entry DEBUG CardService','CardService.java','Mayor','Security Reviewer','✅ Resuelta','S18','LoggingFilter enmascara PAN en todos los logs'],
    ['NC-018-003','Test','Test TC-1603 falta caso 2FA token expirado','CardServiceTest.java','Menor','QA Lead','✅ Resuelta','S18','Caso expiración añadido — token TTL 30s'],
    ['NC-018-004','Documentación','Javadoc faltante CardPinService.changePin()','CardPinService.java','Menor','Code Reviewer','✅ Resuelta','S18','Javadoc completo con @throws SecurityException'],
    ['NC-018-005','Estilo','Constante CARD_LIMIT_DEFAULT debería ser static final','CardLimitsService.java','Menor','Code Reviewer','✅ Resuelta','S18','Movida a constante de clase static final'],
]
for row,d in enumerate(ncs,4):
    for col,val in enumerate(d,1):
        c=ws.cell(row=row,column=col,value=val)
        bg=LGREEN if '✅' in str(val) else (LRED if val in ['Mayor','Crítica'] else WHITE)
        sc(c,bg=bg)
        if col==1: sc(c,bold=True,bg=LGRAY)

for i,w in enumerate([10,14,44,38,10,18,14,10,42],1):
    ws.column_dimensions[get_column_letter(i)].width=w
ws.row_dimensions[1].height=28; ws.row_dimensions[2].height=20
for r in range(3,9): ws.row_dimensions[r].height=28

wb.save(os.path.join(OUT,'NC-Tracker-Sprint18.xlsx'))
print('✅ NC-Tracker-Sprint18.xlsx')

# ── Decision-Log-Sprint18.xlsx ───────────────────────────────────────────────
wb2=openpyxl.Workbook(); ws2=wb2.active
ws2.title='Decision Log Sprint 18'; ws2.sheet_view.showGridLines=False

ws2.merge_cells('A1:H1')
ws2['A1']='DECISION LOG — Sprint 18 · FEAT-016 Gestión de Tarjetas · BankPortal'
ws2['A1'].font=Font(name='Arial',bold=True,size=13,color='FFFFFFFF')
ws2['A1'].fill=hfill(BLUE); ws2['A1'].alignment=ca()
ws2.merge_cells('A2:H2')
ws2['A2']='CMMI Level 3 · Decision Analysis · Total decisiones: 5'
ws2['A2'].font=Font(name='Arial',size=10,color='FF2E5F9E'); ws2['A2'].alignment=ca(); ws2['A2'].fill=hfill(LGRAY)

for col,h in enumerate(['ID','Fecha','Descripción','Alternativas','Decisión','Autor','Impacto','Estado'],1):
    c=ws2.cell(row=3,column=col,value=h); sh(c)

decisions=[
    ['ADR-028','2026-04-23','ShedLock para scheduler multi-instancia','(A) DB lock manual, (B) ShedLock library','B — ShedLock: @SchedulerLock annotation','Tech Lead','R-015-01 cerrado — cero riesgo duplicación','✅ Implementado'],
    ['ARCH-018-01','2026-04-24','PAN masking en logs','(A) Manual en cada log, (B) LoggingFilter global','B — LoggingFilter centralizado con regex PAN','Tech Lead','PCI req.10 satisfecho — código más limpio','✅ Implementado'],
    ['SEC-018-01','2026-04-25','2FA para operaciones sensibles tarjetas','(A) PIN numérico simple, (B) TOTP 2FA obligatorio','B — TOTP 2FA para bloqueo y cambio PIN','Tech Lead + Sec','PCI req.8 satisfecho — seguridad máxima','✅ Implementado'],
    ['DB-018-01','2026-04-26','Tabla card_limits separada vs columnas en cards','(A) Columnas en cards, (B) Tabla card_limits FK','B — Tabla dedicada: extensible y normalizada','Tech Lead','Flyway V18c — separación limpia de dominios','✅ Implementado'],
    ['DIFF-018-01','2026-04-30','Drop columnas plain (V17c) timing','(A) Inmediato S18, (B) Tras validar cifrado','B — Flyway V18b tras confirmar 0 regresión','Tech Lead','V17c cerrada con seguridad — sin rollback riesgo','✅ Implementado'],
]
for row,d in enumerate(decisions,4):
    for col,val in enumerate(d,1):
        c=ws2.cell(row=row,column=col,value=val)
        bg=LGREEN if '✅' in str(val) else WHITE
        sc(c,bg=bg)
        if col==1: sc(c,bold=True,bg=LGRAY,fg=BLUE2)

for i,w in enumerate([12,12,36,44,44,12,44,14],1):
    ws2.column_dimensions[get_column_letter(i)].width=w
ws2.row_dimensions[1].height=28
for r in range(3,9): ws2.row_dimensions[r].height=32

wb2.save(os.path.join(OUT,'Decision-Log-Sprint18.xlsx'))
print('✅ Decision-Log-Sprint18.xlsx')

# ── Quality-Dashboard-Sprint18.xlsx ─────────────────────────────────────────
wb3=openpyxl.Workbook(); ws3=wb3.active
ws3.title='Quality Dashboard'; ws3.sheet_view.showGridLines=False

ws3.merge_cells('A1:F1')
ws3['A1']='QUALITY DASHBOARD — Sprint 18 · FEAT-016 · BankPortal · v1.18.0'
ws3['A1'].font=Font(name='Arial',bold=True,size=14,color='FFFFFFFF')
ws3['A1'].fill=hfill(BLUE); ws3['A1'].alignment=ca()
ws3.merge_cells('A2:F2')
ws3['A2']='CMMI Level 3 · MA SP 1.1/2.1 · PCI-DSS req.3/8/10 · Período: 2026-04-23 → 2026-05-07'
ws3['A2'].font=Font(name='Arial',size=10,color='FF2E5F9E',italic=True)
ws3['A2'].alignment=ca(); ws3['A2'].fill=hfill(LGRAY)

sections=[
    ('MÉTRICAS DE CALIDAD',[
        ('KPI','Valor S18','Valor S17','Umbral','Delta','Estado'),
        ('Tests QA ejecutados','101','45','≥ 30','+56','✅'),
        ('Tests unitarios Developer','16','—','≥ 10','—','✅'),
        ('Tests unitarios acumulados','~677','615','≥ 600','+62','✅'),
        ('Cobertura application','86%','85%','≥ 80%','+1%','✅'),
        ('Defectos en producción','0','0','= 0','=','✅'),
        ('CVEs críticos/altos/medios','0','0','= 0','=','✅'),
        ('NCs bloqueantes CR','0','0','= 0','=','✅'),
        ('Sprint Goal cumplido','100%','100%','≥ 90%','=','✅'),
        ('PCI-DSS req.3','PASS','N/A','PASS','NUEVO','✅'),
        ('PCI-DSS req.8','PASS','N/A','PASS','NUEVO','✅'),
        ('PCI-DSS req.10','PASS','N/A','PASS','NUEVO','✅'),
    ]),
    ('MÉTRICAS DE PROCESO',[
        ('Métrica','Valor','Referencia','Tendencia','',''),
        ('SP entregados','24','24 SP (S17)','→ Estable','',''),
        ('SP acumulados','425','401 (S17)','+24 SP','',''),
        ('Velocidad media (18 sprints)','23.6 SP/sprint','23.6','→ Estable','',''),
        ('Deuda técnica cerrada','4 items (ADR-028/DEBT-026/030/V17c)','3 (S17)','↑ Mejora','',''),
        ('Riesgos cerrados en sprint','3 (R-015-01/R-018-01/02)','2 (S17)','↑ Mejora','',''),
        ('Features completadas','16 (FEAT-001 → 016)','15 (S17)','→ On track','',''),
    ]),
    ('MÉTRICAS ACUMULADAS PROYECTO',[
        ('Métrica','Valor total','','','',''),
        ('Sprints completados','18','','','',''),
        ('SP acumulados','425','','','',''),
        ('Releases producción','v1.1.0 → v1.18.0 (18 releases)','','','',''),
        ('Defectos PRD (histórico)','0','','','',''),
        ('CVEs críticos/altos (histórico)','0','','','',''),
        ('Cobertura actual','86%','','','',''),
        ('CMMI Level activo','3','','','',''),
    ]),
]

cr=4
for st,rows in sections:
    ws3.merge_cells(f'A{cr}:F{cr}')
    ws3[f'A{cr}']=st
    ws3[f'A{cr}'].font=Font(name='Arial',bold=True,size=11,color='FFFFFFFF')
    ws3[f'A{cr}'].fill=hfill(BLUE2); ws3[f'A{cr}'].alignment=ca()
    ws3.row_dimensions[cr].height=20; cr+=1
    for ri,row in enumerate(rows):
        for col,val in enumerate(row,1):
            c=ws3.cell(row=cr,column=col,value=val)
            if ri==0: sh(c,bg=MGRAY,fg='FF1B3A6B')
            else:
                bg=LGREEN if str(val)=='✅' else (LGRAY if col==1 else WHITE)
                sc(c,bold=(col==1),bg=bg)
        cr+=1
    cr+=1

for i,w in enumerate([38,28,24,16,10,10],1):
    ws3.column_dimensions[get_column_letter(i)].width=w

# Hoja Sprint History
wsh=wb3.create_sheet('Sprint History')
wsh.merge_cells('A1:G1')
wsh['A1']='Histórico de Sprints — BankPortal (S1 → S18)'
wsh['A1'].font=Font(name='Arial',bold=True,size=13,color='FFFFFFFF')
wsh['A1'].fill=hfill(BLUE); wsh['A1'].alignment=ca()

for col,h in enumerate(['Sprint','Feature','SP','SP Acum.','Cobertura','Tests Acum.','Defectos PRD'],1):
    c=wsh.cell(row=2,column=col,value=h); sh(c)

hist=[
    ['S1','FEAT-001 Auth',18,18,'72%',88,0],
    ['S2','Deuda S1',11,29,'74%',104,0],
    ['S3','FEAT-002 Dashboard',21,50,'75%',142,0],
    ['S4','FEAT-003 Transferencias',22,72,'76%',178,0],
    ['S5-8','FEAT-004/005',23,95,'78%',230,0],
    ['S9','Deuda técnica',11,106,'79%',280,0],
    ['S10','FEAT-008 Seguridad',24,130,'80%',320,0],
    ['S11','FEAT-009 Reporting',22,152,'81%',355,0],
    ['S12','FEAT-010 Core Banking',23,175,'82%',390,0],
    ['S13','FEAT-011 Multi-cuenta',24,199,'83%',430,0],
    ['S14','FEAT-012 OAuth2/PKCE',22,221,'83%',468,0],
    ['S15','FEAT-013 Auditoría',22,243,'84%',500,0],
    ['S16','FEAT-014 Push VAPID',24,267,'84%',553,0],
    ['S17','FEAT-015 Trans. Programadas',24,291,'85%',615,0],
    ['S18','FEAT-016 Gestión Tarjetas',24,315,'86%',677,0],
]
for row,d in enumerate(hist,3):
    for col,val in enumerate(d,1):
        c=wsh.cell(row=row,column=col,value=val)
        sc(c,bold=(col==1),bg=LGRAY if col==1 else (LGREEN if col==7 and val==0 else WHITE))

for i,w in enumerate([8,36,8,10,12,12,14],1):
    wsh.column_dimensions[get_column_letter(i)].width=w

wb3.save(os.path.join(OUT,'Quality-Dashboard-Sprint18.xlsx'))
print('✅ Quality-Dashboard-Sprint18.xlsx')

print('\n✅ Sprint 18 — 3 ficheros Excel generados en:', OUT)
