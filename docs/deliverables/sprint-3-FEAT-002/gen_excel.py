# SOFIA Documentation Agent — gen_excel.py — Sprint 3 — FEAT-002 Sesiones Activas
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from pathlib import Path

EXCEL_DIR = Path(__file__).parent / 'excel'
EXCEL_DIR.mkdir(exist_ok=True)

BLUE='1B3A6B'; MED='2E5F9E'; VL='EBF3FB'; GPASS='C6EFCE'; GDONE='E2EFDA'
RED='FFCCCC'; YEL='FFEB9C'; WHITE='FFFFFF'; GRAY='CCCCCC'
SPRINT='3'; VERSION='v1.2.0'; FECHA='2026-04-25'; FEAT='FEAT-002'

def side(): return Side(style='thin', color=GRAY)
def brd(): return Border(left=side(),right=side(),top=side(),bottom=side())
def hfont(): return Font(name='Arial',bold=True,color=WHITE,size=11)
def dfont(bold=False): return Font(name='Arial',size=10,bold=bold)
def hfill(): return PatternFill('solid',fgColor=BLUE)
def afill(): return PatternFill('solid',fgColor=VL)
def wfill(): return PatternFill('solid',fgColor=WHITE)
def sfill(s):
    u=str(s).upper()
    if any(k in u for k in ('PASS','OK','DONE','APPROVED','CUMPLE','CERRADO')): return PatternFill('solid',fgColor=GPASS)
    if any(k in u for k in ('FAIL','ERROR')): return PatternFill('solid',fgColor=RED)
    if any(k in u for k in ('BLOCKED','OPEN','WARN','ABIERTO','PARCIAL')): return PatternFill('solid',fgColor=YEL)
    return wfill()
def cc(): return Alignment(horizontal='center',vertical='center',wrap_text=True)
def lc(): return Alignment(horizontal='left',vertical='top',wrap_text=True)
def ap(cell,font=None,fill=None,align=None):
    if font: cell.font=font
    if fill: cell.fill=fill
    if align: cell.alignment=align
    cell.border=brd()
def title(ws,row,text,ncols):
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=ncols)
    c=ws.cell(row=row,column=1,value=text)
    c.font=Font(name='Arial',bold=True,size=13,color=WHITE); c.fill=hfill(); c.alignment=cc()
    ws.row_dimensions[row].height=26
def hrow(ws,row,cols):
    for i,(text,width) in enumerate(cols,1):
        c=ws.cell(row=row,column=i,value=text)
        ap(c,hfont(),hfill(),cc()); ws.column_dimensions[get_column_letter(i)].width=width
    ws.row_dimensions[row].height=22
def drow(ws,row,vals,alt=False,scols=None):
    f=afill() if alt else wfill()
    for i,v in enumerate(vals,1):
        c=ws.cell(row=row,column=i,value=v)
        use=sfill(v) if scols and i in scols else f
        ap(c,dfont(),use,lc()); ws.row_dimensions[row].height=18

def build_quality_dashboard():
    wb=Workbook(); ws=wb.active; ws.title='Resumen'
    title(ws,1,f'Quality Dashboard — Sprint {SPRINT} — BankPortal — {FEAT}',5)
    ws.merge_cells('A2:E2'); c=ws.cell(row=2,column=1,value=f'{FEAT} | SOFIA QA Agent | {FECHA} | {VERSION}')
    c.font=Font(name='Arial',size=10,italic=True,color='444444'); c.fill=PatternFill('solid',fgColor=VL); c.alignment=cc()
    hrow(ws,3,[('Área',26),('TCs Plan',12),('TCs Real',12),('Cobertura',12),('Estado',12)])
    areas=[
        ('US-101 — Listar sesiones',          6, 6,'100%','PASS'),
        ('US-102 — Revocar sesión + OTP',      8, 8,'100%','PASS'),
        ('US-103 — Timeout configurable',      5, 5,'100%','PASS'),
        ('US-104 — LRU concurrencia',          7, 7,'100%','PASS'),
        ('US-105 — Alertas email',             6, 6,'100%','PASS'),
        ('DEBT-003 — POST /deactivate',        3, 3,'100%','PASS'),
        ('Seguridad + integración',            5, 5,'100%','PASS'),
    ]
    for i,(a,p,r,cov,st) in enumerate(areas): drow(ws,i+4,[a,p,r,cov,st],i%2==1,{5})
    tr=len(areas)+4
    for col,val,fill in [(1,'TOTAL Sprint 3',BLUE),(2,'40',GPASS),(3,'40',GPASS),(4,'100%',GPASS),(5,'PASS',GPASS)]:
        c=ws.cell(row=tr,column=col,value=val)
        c.font=Font(name='Arial',bold=True,size=11,color=WHITE if col==1 else '000000')
        c.fill=PatternFill('solid',fgColor=fill); c.alignment=cc(); c.border=brd()
    kr=tr+2; ws.merge_cells(start_row=kr,start_column=1,end_row=kr,end_column=5)
    c=ws.cell(row=kr,column=1,value='KPIs Acumulados'); c.font=hfont(); c.fill=hfill(); c.alignment=cc()
    kpis=[('TCs acumulados FEAT-001+002','137'),('Defectos totales','0'),('PCI-DSS req. 8.2.4+8.2.8','CUMPLE'),('Gates HITL acumulados','15'),('Velocidad media','24 SP/sprint')]
    for j,(k,v) in enumerate(kpis):
        r=kr+1+j; c1=ws.cell(row=r,column=1,value=k); ap(c1,dfont(True),afill() if j%2==0 else wfill(),lc())
        ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=3)
        c2=ws.cell(row=r,column=4,value=v); ap(c2,Font(name='Arial',bold=True,size=11),PatternFill('solid',fgColor=GPASS),cc())
        ws.merge_cells(start_row=r,start_column=4,end_row=r,end_column=5)

    ws2=wb.create_sheet('Ejecucion TCs')
    title(ws2,1,f'Ejecución Casos de Prueba — Sprint {SPRINT}',7)
    hrow(ws2,2,[('TC ID',14),('US/DEBT',10),('Descripción',40),('Nivel',16),('Tipo',14),('Prior.',10),('Estado',12)])
    tcs=[
        ('TC-101-001','US-101','Lista sesiones activas con device info','Funcional','Happy Path','Alta','PASS'),
        ('TC-101-002','US-101','Sin sesiones activas — lista vacía','Funcional','Edge Case','Media','PASS'),
        ('TC-101-003','US-101','Token expirado incluido en lista','Funcional','Edge Case','Alta','PASS'),
        ('TC-101-004','US-101','Sin JWT — 401','Seguridad','Seguridad','Alta','PASS'),
        ('TC-101-005','US-101','Solo sesiones propias (userId claim)','Seguridad','Aislamiento','Alta','PASS'),
        ('TC-101-006','US-101','Sesión actual marcada con isCurrent=true','Funcional','Verificación','Media','PASS'),
        ('TC-102-001','US-102','Revocar sesión con OTP correcto','Funcional','Happy Path','Alta','PASS'),
        ('TC-102-002','US-102','OTP incorrecto — revocación rechazada','Funcional','Error Path','Alta','PASS'),
        ('TC-102-003','US-102','Revocar sesión de otro usuario — 403','Seguridad','Aislamiento','Alta','PASS'),
        ('TC-102-004','US-102','Sesión ya revocada — idempotente','Funcional','Edge Case','Media','PASS'),
        ('TC-102-005','US-102','Revocar all — sesión actual excluida','Funcional','Happy Path','Alta','PASS'),
        ('TC-102-006','US-102','audit_log registra SESSION_REVOKED','Auditoría','Trazabilidad','Alta','PASS'),
        ('TC-102-007','US-102','SSE notifica sesión revocada','Integración','Integración','Media','PASS'),
        ('TC-102-008','US-102','JWT revocado en Redis blacklist','Seguridad','Seguridad','Alta','PASS'),
        ('TC-103-001','US-103','Timeout 15 min — sesión expira correctamente','Funcional','Happy Path','Alta','PASS'),
        ('TC-103-002','US-103','Timeout 60 min persiste entre reinicios','Funcional','Persistencia','Alta','PASS'),
        ('TC-103-003','US-103','Cambio timeout audita TIMEOUT_CHANGED','Auditoría','Trazabilidad','Media','PASS'),
        ('TC-103-004','US-103','Valor inválido (0 min) → 400','Funcional','Error Path','Alta','PASS'),
        ('TC-103-005','US-103','Timeout aplicado en filtro JWT','Integración','Integración','Alta','PASS'),
        ('TC-104-001','US-104','4a sesión → sesión más antigua revocada LRU','Funcional','Happy Path','Alta','PASS'),
        ('TC-104-002','US-104','Email enviado al revocar sesión LRU','Integración','Notificación','Media','PASS'),
        ('TC-104-003','US-104','3 sesiones → no LRU eviction','Funcional','Edge Case','Media','PASS'),
        ('TC-104-004','US-104','Redis falla → LRU en memoria','Resilencia','Fallback','Alta','PASS'),
        ('TC-104-005','US-104','audit_log registra SESSION_EVICTED_LRU','Auditoría','Trazabilidad','Alta','PASS'),
        ('TC-104-006','US-104','LRU considera lastActivityAt','Funcional','Verificación','Alta','PASS'),
        ('TC-104-007','US-104','Concurrencia — race condition segura','Técnica','Thread-safety','Alta','PASS'),
        ('TC-105-001','US-105','Email enviado en login nuevo dispositivo','Funcional','Happy Path','Alta','PASS'),
        ('TC-105-002','US-105','Preferencia desactivada → sin email','Funcional','Configuración','Media','PASS'),
        ('TC-105-003','US-105','Email incluye IP y dispositivo','Funcional','Verificación','Media','PASS'),
        ('TC-105-004','US-105','SMTP falla — evento en dead letter queue','Resilencia','Error Handling','Alta','PASS'),
        ('TC-105-005','US-105','Login mismo dispositivo → sin email','Funcional','Edge Case','Media','PASS'),
        ('TC-105-006','US-105','audit_log registra NOTIFICATION_SENT','Auditoría','Trazabilidad','Media','PASS'),
        ('TC-D003-001','DEBT-003','POST /deactivate funciona — DELETE deprecado','Funcional','Migración','Alta','PASS'),
        ('TC-D003-002','DEBT-003','DELETE /deactivate → 405 Method Not Allowed','Funcional','Retrocompat.','Alta','PASS'),
        ('TC-D003-003','DEBT-003','Documentación OpenAPI actualizada','Documental','Verificación','Media','PASS'),
        ('TC-SEG-S3-001','Seguridad','Session token no predecible (UUID v4)','Seguridad','OWASP','Alta','PASS'),
        ('TC-SEG-S3-002','Seguridad','JWT revocado no reutilizable','Seguridad','Seguridad','Alta','PASS'),
        ('TC-SEG-S3-003','Seguridad','IP subnet enmascarada en respuesta','Privacidad','GDPR','Media','PASS'),
        ('TC-INT-S3-001','Integración','Flujo completo login→list→revoke→verify','E2E','Integración','Alta','PASS'),
        ('TC-INT-S3-002','Integración','Timeout + LRU interacción correcta','E2E','Integración','Alta','PASS'),
    ]
    for i,tc in enumerate(tcs): drow(ws2,i+3,list(tc),i%2==1,{7})

    ws3=wb.create_sheet('PCI-DSS 4.0')
    title(ws3,1,f'Checklist PCI-DSS 4.0 — Sprint {SPRINT}',4)
    hrow(ws3,2,[('Requisito',12),('Descripción',42),('Verificación',32),('Estado',12)])
    pci=[
        ('8.2.4','Sesiones de usuarios autenticados deben poder cerrarse','Revocación individual + bulk con audit trail','CUMPLE'),
        ('8.2.8','Timeout de inactividad para sesiones','Timeout configurable 15/30/60 min en JwtFilter','CUMPLE'),
        ('8.2.6','Historial de sesiones activas visible','GET /sessions lista con device + IP + lastActivity','CUMPLE'),
        ('10.2.4','Log de inicio y fin de sesión','audit_log SESSION_CREATED + SESSION_REVOKED + timestamp','CUMPLE'),
    ]
    for i,row in enumerate(pci): drow(ws3,i+3,list(row),i%2==1,{4})

    ws4=wb.create_sheet('Métricas')
    title(ws4,1,f'Métricas de Calidad — Sprint {SPRINT}',5)
    hrow(ws4,2,[('Métrica',30),('Valor',15),('Umbral',15),('Estado',12),('Notas',25)])
    mets=[
        ('TCs ejecutados','40/40','100%','PASS',''),
        ('Defectos críticos','0','0','PASS',''),
        ('Cobertura unitaria','~89%','>=80%','PASS',''),
        ('Tests E2E integración','8/8','>=80%','PASS',''),
        ('PCI-DSS req. 8.2.4/8.2.8','CUMPLE','100%','PASS',''),
    ]
    for i,row in enumerate(mets): drow(ws4,i+3,list(row),i%2==1,{4})
    wb.save(EXCEL_DIR/f'Quality-Dashboard-Sprint{SPRINT}.xlsx')
    print('  ✅',EXCEL_DIR/f'Quality-Dashboard-Sprint{SPRINT}.xlsx')

def build_sprint_metrics():
    wb=Workbook(); ws=wb.active; ws.title='Resumen Sprint'
    title(ws,1,f'Sprint Metrics — Sprint {SPRINT} — {FEAT}',5)
    ws.merge_cells('A2:E2'); c=ws.cell(row=2,column=1,value=f'Sprint Goal ALCANZADO | 2026-04-14 - {FECHA} | {VERSION}')
    c.font=Font(name='Arial',size=10,italic=True); c.fill=PatternFill('solid',fgColor=VL); c.alignment=cc()
    hrow(ws,3,[('Métrica',28),('Planificado',14),('Real',14),('Variación',14),('Estado',14)])
    items=[('Story Points',24,24,'0','OK'),('US completadas',5,5,'0','OK'),('DEBT resueltas',1,1,'0','OK'),('Tests nuevos','-',40,'-','OK'),('Defectos QA',0,0,'0','OK'),('Gates HITL',5,5,'0','OK')]
    for i,(m,p,r,v,st) in enumerate(items):
        row=i+4
        c1=ws.cell(row=row,column=1,value=m); ap(c1,dfont(True),afill() if i%2==1 else wfill(),lc())
        for col,val in [(2,p),(3,r),(4,v)]:
            c=ws.cell(row=row,column=col,value=val); ap(c,dfont(),afill() if i%2==1 else wfill(),cc())
        cs=ws.cell(row=row,column=5,value=st); ap(cs,Font(name='Arial',size=10,bold=True),sfill(st),cc())
    sr=len(items)+4; ws.merge_cells(start_row=sr,start_column=1,end_row=sr,end_column=3)
    c=ws.cell(row=sr,column=1,value=f'VELOCIDAD SPRINT {SPRINT}'); c.font=Font(name='Arial',bold=True,size=12,color=WHITE); c.fill=hfill(); c.alignment=cc()
    ws.merge_cells(start_row=sr,start_column=4,end_row=sr,end_column=5)
    c2=ws.cell(row=sr,column=4,value='24 SP / sprint'); c2.font=Font(name='Arial',bold=True,size=12); c2.fill=PatternFill('solid',fgColor=GPASS); c2.alignment=cc(); c2.border=brd()

    ws2=wb.create_sheet('Items Detalle')
    title(ws2,1,f'Items Sprint {SPRINT} — Detalle',5)
    hrow(ws2,2,[('ID',12),('Título',42),('SP',8),('Tipo',14),('Estado',12)])
    det=[('US-101','Listar sesiones activas del usuario',5,'new-feature','DONE'),('US-102','Revocar sesión individual con confirmación OTP',4,'new-feature','DONE'),('US-103','Timeout de inactividad configurable',3,'new-feature','DONE'),('US-104','Control de sesiones concurrentes LRU',5,'new-feature','DONE'),('US-105','Notificaciones email login nuevo dispositivo',4,'new-feature','DONE'),('DEBT-003','POST /deactivate migración REST',1,'tech-debt','DONE'),('ACT-Flyway','Flyway V5 session_tokens + session_metadata',2,'infra','DONE')]
    for i,row in enumerate(det): drow(ws2,i+3,list(row),i%2==1,{5})
    for col,val,fill in [(1,'TOTAL',BLUE),(2,'7 items',BLUE),(3,24,BLUE),(4,'',BLUE),(5,'24/24 SP',GPASS)]:
        c=ws2.cell(row=len(det)+3,column=col,value=val)
        c.font=Font(name='Arial',bold=True,size=11,color=WHITE if fill==BLUE else '000000')
        c.fill=PatternFill('solid',fgColor=fill); c.alignment=cc(); c.border=brd()
    wb.save(EXCEL_DIR/f'Sprint-Metrics-Sprint{SPRINT}.xlsx')
    print('  ✅',EXCEL_DIR/f'Sprint-Metrics-Sprint{SPRINT}.xlsx')

def build_velocity_report():
    wb=Workbook(); ws=wb.active; ws.title='Velocidad'
    title(ws,1,'Velocity Report — BankPortal — Acumulado',6)
    ws.merge_cells('A2:F2'); c=ws.cell(row=2,column=1,value='BankPortal · Banco Meridian · 3 Sprints')
    c.font=Font(name='Arial',size=10,italic=True); c.fill=PatternFill('solid',fgColor=VL); c.alignment=cc()
    hrow(ws,3,[('Sprint',12),('Período',22),('SP Plan',14),('SP Real',14),('Velocidad',14),('Variación',14)])
    sprints=[('Sprint 1','2026-03-02 — 2026-03-16',24,24),('Sprint 2','2026-03-30 — 2026-04-10',24,24),('Sprint 3','2026-04-14 — 2026-04-25',24,24)]
    for i,(sp,per,plan,real) in enumerate(sprints):
        row=i+4
        for col,val,aln in [(1,sp,cc()),(2,per,lc()),(3,plan,cc()),(4,real,cc())]:
            c=ws.cell(row=row,column=col,value=val); ap(c,dfont(col==1),afill() if i%2==1 else wfill(),aln)
        cv=ws.cell(row=row,column=5,value=real); ap(cv,Font(name='Arial',bold=True,size=10),PatternFill('solid',fgColor=GPASS),cc())
        cvar=ws.cell(row=row,column=6,value=real-plan); ap(cvar,dfont(),afill() if i%2==1 else wfill(),cc())
    ar=len(sprints)+4; ws.merge_cells(start_row=ar,start_column=1,end_row=ar,end_column=2)
    c=ws.cell(row=ar,column=1,value='VELOCIDAD MEDIA'); c.font=Font(name='Arial',bold=True,color=WHITE); c.fill=hfill(); c.alignment=cc(); c.border=brd()
    for col,val in [(3,24),(4,24),(5,24),(6,0)]:
        c=ws.cell(row=ar,column=col,value=val); ap(c,Font(name='Arial',bold=True),PatternFill('solid',fgColor=GPASS),cc())
    chart=BarChart(); chart.type='col'; chart.title='Velocidad por Sprint'; chart.y_axis.title='SP'; chart.style=10; chart.width=18; chart.height=10
    chart.add_data(Reference(ws,min_col=4,min_row=3,max_row=ar-1),titles_from_data=True)
    chart.set_categories(Reference(ws,min_col=1,min_row=4,max_row=ar-1)); ws.add_chart(chart,'H3')
    wb.save(EXCEL_DIR/f'Velocity-Report-Sprint{SPRINT}.xlsx')
    print('  ✅',EXCEL_DIR/f'Velocity-Report-Sprint{SPRINT}.xlsx')

if __name__=='__main__':
    print(f'\n📊 SOFIA Documentation Agent — Sprint {SPRINT} Excel docs...')
    build_quality_dashboard(); build_sprint_metrics(); build_velocity_report()
    print('✅ Excel docs Sprint 3 completados\n')
